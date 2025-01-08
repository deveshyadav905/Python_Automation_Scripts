import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from category_finder import category_finder

class open_url_browser:
    def __init__(self, sheet_path):
        self.sheet_path = sheet_path
    
    def open_url(self):
        # chromedriver_path = "/home/roopchand/Downloads/chromedriver-linux64/chromedriver"
        # Create ChromeOptions object
        df=pd.read_excel(self.sheet_path, sheet_name='Data')
        data=df['check_URL'].to_list()
        chrome_options = Options()
        # chrome_options.add_argument(f"webdriver.chrome.driver={chromedriver_path}")
        with webdriver.Chrome(options=chrome_options) as driver:    
            for url in data:
                try:
                    driver.execute_script(f"window.open('{url.strip()}');")
                    # time.sleep(1)
                except:
                    print(f"fail this url {url}")    

            int(input("Inter any number to get brower url "))
            self.brower_to_sheet(driver)
            driver.quit()
    
    def brower_to_sheet(self, driver):
        tab_handles = driver.window_handles
        urls = []
        for handle in tab_handles:
            driver.switch_to.window(handle)
            urls.append(driver.current_url)
        wb = openpyxl.load_workbook(self.sheet_path)
        if 'url1' not in wb.sheetnames:
            wb.create_sheet('url1')
        sheet = wb['url1']
        # Add column name
        sheet['A1'] = "source_url"
        for i, url in enumerate(urls, start=2):  # Start from row 2 to leave the first row for the column name
            sheet.cell(row=i, column=1, value=url)
        wb.save(self.sheet_path)
        print(f"save url to sheet succesfully!!")
        value=input("Enter 'Y' for find category of url else Enter 'N' ")
        if value == 'y' or value == "Y":
            category_finder()
        else:
            pass    

sheet_path = "/home/charanjeet/projects/newsdatafeeds/feeds_crawler/open_url_browser.xlsx"
obj=open_url_browser(sheet_path)
obj.open_url()